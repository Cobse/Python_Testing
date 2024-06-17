import openpyxl
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont

wb = openpyxl.Workbook()
ws = wb.active

fonts = [
    InlineFont(sz=10, color="FF0000"),
    InlineFont(sz=20),
    InlineFont(sz=14),
]

text_strings = ["Anker", "Betty", "May"]

text_blocks = []
for i, text in enumerate(text_strings):
    font = fonts[i % len(fonts)]  # Loop through the fonts repeatedly if more text strings than fonts
    text_block = TextBlock(font=font, text=text)
    text_blocks.append(text_block)
print(text_blocks)
ws["A1"] = CellRichText(text_blocks)

# Define the range you want to loop through
start_row = 1
end_row = 10
start_column = 1
end_column = 3

# Loop through the range and write values to cells
for row in range(start_row, end_row + 1):
    for column in range(start_column, end_column + 1):
        cell_value = CellRichText(text_blocks)
        ws.cell(row=row, column=column).value = cell_value


wb.save("output.xlsx")